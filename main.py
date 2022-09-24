"""Create the window and widgets"""

from datetime import datetime
from pathlib import Path
import csv
import os
import subprocess
import tkinter as tk
import tkinter.filedialog
import tkinter.messagebox
import openpyxl

# List Of Checkbox Items
copy_cases = ["Copy Renders", "Copy Source Images"]
check = []

# Creating object of tk class
window = tk.Tk()

# Setting the title and background color
# disabling the resizing property
window.geometry("600x320")
window.title("Saffronic Packing Tool")
window.config(background="white")

# Creating tkinter variable
sourceLocation = tk.StringVar()
destinationLocation = tk.StringVar()
# Store created directories for rollback
created_directories = []


def source_browse():
    """Opening the file-dialog prompting the user to
    select a .xlsx file"""
    window.sourceText.delete(0, tk.END)
    window.file_list = list(
        tkinter.filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx")])
    )
    window.sourceText.insert("0", window.file_list)


def destination_browse():
    """Opening the file-dialog directory prompting
    the user to select destination folder to
    which files are to be copied using the
    tkinter.filedialog.askopendirectory() method.
    Setting initialdir argument is optional"""
    window.destinationText.delete(0, tk.END)
    destinationdirectory = tkinter.filedialog.askdirectory()

    # Displaying the selected directory in the
    # window.destinationText tk.Entry using
    # window.destinationText.insert()
    # window.destinationText.insert("1", destinationdirectory)
    window.destinationText.insert("0", destinationdirectory)
    if(len(window.file_list) > 0 and len(destinationdirectory) > 0):
        window.copyButton["state"] = tk.NORMAL


def log_result(src, destination_location, result, operation_type="copy"):
    """Log results of all operations in a csv file"""
    
    log_dir = Path(window.file_list[0]).parent
    csv_path= f"{log_dir}\\.saffronic_log.csv"
    # if file does not exist, create it

    if Path(csv_path).is_file() is False:
        with open(csv_path, "w", encoding="utf-8") as log_file:
            csv_writer = csv.writer(log_file)
            csv_writer.writerow(
                [
                    "Timestamp",
                    "Operation Type",
                    "Source",
                    "Destination",
                    "Return code",
                    "Stdout",
                    "Stderr",
                ]
            )

    # open a csv file in the directory where spreadsheet is located
    with open(
        csv_path, "a", encoding="utf-8", newline=""
    ) as log_file:
        csv_writer = csv.writer(log_file)
        csv_writer.writerow(
            [
                datetime.now().isoformat(),
                operation_type,
                str(src),
                str(destination_location),
                result.returncode,
                result.stdout,
                result.stderr,
            ]
        )


def copy_file():
    """Retrieving the source file selected by the
    user in the SourceBrowse() and storing it in a
    variable named file_list Retrieving the destination location from the
    textvariable using destinationLocation.get() and
    storing in destination_location
    """
    file_list = window.file_list
    dataframe = openpyxl.load_workbook(file_list[0])  # reads only first excel file
    dataframe1 = dataframe.active

    for i in range(2, dataframe1.max_row + 1):
        for j in range(1, dataframe1.max_column + 1):
            # check for the checkbuttons
            if check[j - 2].get() == 1 or j == 1:
                src = dataframe1.cell(row=i, column=j).value
                destination_location = Path(destinationLocation.get())
                if src is None:
                    continue
                src = Path(src)
                folder_substructure = src.parent
                folder_substructure = list(folder_substructure.parts[1:])
                # keep note of all directories created
                created_directories.append(folder_substructure[0])
                destination_location = Path(
                    os.path.join(destination_location, *folder_substructure)
                )
                # pylint: disable=W1510
                result = subprocess.run(
                    [
                        "xcopy",
                        f"{str(src)}",
                        f"{str(destination_location)}\\",
                        "/H/C/I/Y",
                    ],
                    # check=True,
                    capture_output=True,
                    text=True,
                )
                log_result(src, destination_location, result)

    window.rollbackButton["state"] = tk.NORMAL
    tkinter.messagebox.showinfo("Done", "Operation Completed")


def rollback():
    """Rollback copy operation"""
    for directory_new in created_directories:
        destination_location = Path(destinationLocation.get())
        destination_location = destination_location.joinpath(directory_new)
        # pylint: disable=W1510
        # /S to delete subdirectories and /Q to suppress confirmation
        result = subprocess.run(
            ["rmdir", f"{destination_location}", "/S/Q"],
            # check=True,
            capture_output=True,
            text=True,
            shell=True,
        )
        log_dir = Path(window.file_list[0]).parent
        csv_path= f"{log_dir}\\.saffronic_log.csv"

        # just try to execute an action (eg.rename) on the file
        # if it is open and running, it will not work
        if os.path.exists(csv_path):
            try:
                os.rename(csv_path, csv_path)
            except OSError:
                message=f"Access-error on file {csv_path} !, Try closing the file"
                tkinter.messagebox.showinfo("Error", message)

        log_result(
                    destination_location,
                    destination_location,
                    result,
                    operation_type="rollback",
                )
    tkinter.messagebox.showinfo("Done", "Operation Reverted")

# Create widgets
source_browseButton = tk.Button(
    window, text="Input Path", command=source_browse, width=15, bg="#90ee90", fg="black"
)
source_browseButton.grid(row=2, column=0, pady=15, padx=15)

window.sourceText = tk.Entry(window, width=50, textvariable=sourceLocation)
window.sourceText.grid(row=2, column=1, pady=15, padx=15, columnspan=2)

dest_browseButton = tk.Button(
    window,
    text="Destination Path",
    command=destination_browse,
    width=15,
    bg="#90ee90",
    fg="black",
)
dest_browseButton.grid(row=3, column=0, pady=15, padx=15)

window.destinationText = tk.Entry(window, width=50, textvariable=destinationLocation)
window.destinationText.grid(row=3, column=1, pady=15, padx=15, columnspan=2)

window.copyButton = tk.Button(
    window,
    text="Submit",
    command=copy_file,
    width=15,
    bg="#8FBC8F",
    fg="black",
    state=tk.DISABLED,
)
window.copyButton.grid(row=7, column=1, pady=5, padx=1)

window.rollbackButton = tk.Button(
    window,
    text="Undo",
    command=rollback,
    width=15,
    bg="#FFCCCB",
    fg="black",
    state=tk.DISABLED,
)
window.rollbackButton.grid(row=7, column=2, pady=5, padx=1)


for x, n in enumerate(copy_cases):
    check.append(tk.IntVar(window, value=0))
    cb = tk.Checkbutton(window, text=n, variable=check[x])
    cb.grid(row=x + 4, column=0, sticky=tk.W, padx=10, pady=10)


# Defining infinite loop
window.mainloop()
