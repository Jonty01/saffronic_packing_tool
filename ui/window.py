"""Create the window and widgets"""

from subprocess import call
from tkinter import END, Tk, filedialog, messagebox, StringVar, IntVar, Button, Entry, Checkbutton, W
import openpyxl

# List Of Checkbox Items
copy_cases = ["Copy Renders", "Copy Source Images"]
check = []

# Creating object of tk class
window = Tk()

# Setting the title and background color
# disabling the resizing property
window.geometry("830x320")
window.title("Saffronic Packing Tool")
window.config(background="white")

# Creating tkinter variable
sourceLocation = StringVar()
destinationLocation = StringVar()


def source_browse():
    window.sourceText.delete(0, END)
    window.file_list=list(filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx")]))
    window.sourceText.insert('0', window.file_list)

def destination_browse():
    """Opening the file-dialog directory prompting
    the user to select destination folder to
    which files are to be copied using the
    filedialog.askopendirectory() method.
    Setting initialdir argument is optional"""
    window.destinationText.delete(0, END)
    destinationdirectory = filedialog.askdirectory()

    # Displaying the selected directory in the
    # window.destinationText Entry using
    # window.destinationText.insert()
    # window.destinationText.insert("1", destinationdirectory)
    window.destinationText.insert("0", destinationdirectory)


def copy_file():
    """Retrieving the source file selected by the
    user in the SourceBrowse() and storing it in a
    variable named file_list Retrieving the destination location from the
    textvariable using destinationLocation.get() and
    storing in destination_location
    """
    file_list = window.file_list
    dataframe=openpyxl.load_workbook(file_list[0])  #reads only first excel file
    dataframe1=dataframe.active


    for i in range(2, dataframe1.max_row+1):
        for j in range(1, dataframe1.max_column+1):
            if check[j-2].get()==1 or j==1:
                src=dataframe1.cell(row=i, column=j).value
                destination_location = destinationLocation.get()
                if src is None:
                    continue
                src=src.replace("/", "\\")
                last_backslash=src.rindex("\\")
                first_backslash=src.find("\\")
                folder_substructure=src[first_backslash:last_backslash]
                destination_location=destination_location.replace("/", "\\")  #D:\\test1\\YUvraj
                destination_location=destination_location+"\\"+folder_substructure+"\\"    # D:\\test1\\YUvraj\\
                print(src)
                print(destination_location)
                call (["xcopy", f"{src}", f"{destination_location}", "/H/C/I/Y"])
                # shutil.copy(f, destination_location)

    # check in one loop
    messagebox.showinfo("Done", "Operation Completed")


# Create widgets
source_browseButton = Button(
    window, text="Input Path", command=source_browse, width=15, bg="#90ee90", fg="black"
)
source_browseButton.grid(row=2, column=0, pady=15, padx=15)

window.sourceText = Entry(window, width=50, textvariable=sourceLocation)
window.sourceText.grid(row=2, column=1, pady=15, padx=15, columnspan=2)

dest_browseButton = Button(
    window,
    text="Destination Path",
    command=destination_browse,
    width=15,
    bg="#90ee90",
    fg="black",
)
dest_browseButton.grid(row=3, column=0, pady=15, padx=15)

window.destinationText = Entry(window, width=50, textvariable=destinationLocation)
window.destinationText.grid(row=3, column=1, pady=15, padx=15, columnspan=2)

window.copyButton = Button(
    window, text="Submit", command=copy_file, width=15, bg="#8FBC8F", fg="black"
)
window.copyButton.grid(row=7, column=1, pady=5, padx=5)

window.quitButton = Button(
    window, text="Quit", command=exit, width=15, bg="#FFCCCB", fg="black"
)
window.quitButton.grid(row=7, column=2, pady=5, padx=5)

for x, n in enumerate(copy_cases):
    check.append(IntVar(window, value=0))
    cb = Checkbutton(window, text=n, variable=check[x])
    cb.grid(row=x + 4, column=0, sticky=W, padx=10, pady=10)


# Defining infinite loop
window.mainloop()
